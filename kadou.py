import openpyxl
import datetime

# 所属PJが変わるたびに変更
project_name = "サンプルプロジェクト"

# project_name = input("案件名を入力してください。")
current_month = input("何月分ですか？")

# 開始時間（9:30）
start_time_h = 9
start_time_m = 30

# 休憩時間（1時間）
rest_time_h = 1

start = datetime.time(start_time_h, start_time_m, 00)
end = datetime.time(20, 00, 00)
rest = datetime.time(hour=rest_time_h)

# 勤務時間
wrk_hours = (datetime.datetime.combine(datetime.date.today(), end) \
    - datetime.timedelta(hours=start_time_h, minutes=start_time_m) \
    - datetime.timedelta(hours=rest_time_h)).strftime("%H:%M:%S")

file_name = "稼働報告{}月.xlsx".format(current_month)

print(file_name)

# excelファイルを開く
#with open(file_name, "w", encoding="utf_8_sig")

wb = openpyxl.load_workbook(file_name)
sheet = wb.worksheets[0]

for dt_row in range(5, 36):
    if sheet.cell(dt_row, 3).value != "土" and sheet.cell(dt_row, 3).value != "日":
        if sheet.cell(dt_row, 3).value != None:
            sheet.cell(dt_row, 4).value = project_name
            sheet.cell(dt_row, 5).value = start
            sheet.cell(dt_row, 6).value = end
            sheet.cell(dt_row, 7).value = wrk_hours
            sheet.cell(dt_row, 8).value = rest

wb.save(file_name)
wb.close()

print("処理が終了しました。")



